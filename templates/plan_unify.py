# -*- coding: utf-8 -*-
"""
생산계획 통일 양식 생성기 / 기존 담당자 파일 변환기

사용법
  python plan_unify.py template  <출력.xlsx>                  # 빈 통일 양식(작성기준 시트 포함)
  python plan_unify.py convert   <출력.xlsx> <원본1.xlsx> ...  # 담당자별 파일을 통일 양식으로 변환

원본 양식 자동 판별
  - 치약형   : A열 시작, 시트명 '9.4(금)', 헤더 2행
  - 염모/튜브형: B열 시작, 시트명 '09월03일 염모', 헤더 2행, 하단 '지연품목' 블록(튜브)
  - 멀티/직선형: B열 시작, 시트명 '9.4금(멀)', 헤더 3행, 상단 '지연오더' 블록, 제품명 끝에 ' / 8/20오더'
"""
import re, sys, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

YEAR = 2026
FONT = '맑은 고딕'
NAVY = '0F2438'; GREEN = '5B7F52'
HDR_FILL = PatternFill('solid', fgColor='DCE5D8')
FORMULA_FILL = PatternFill('solid', fgColor='F2F2F2')
DELAY_FILL = PatternFill('solid', fgColor='FFF2CC')
FOOT_FILL = PatternFill('solid', fgColor='F7F7F7')
thin = Side(style='thin', color='999999'); med = Side(style='medium', color='555555')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WD = '월화수목금토일'

# 통일 열 정의 (열, 헤더, 너비, 입력/수식, 설명)
COLS = [
    ('A', '구분',     6,  '입력', '정상 / 지연 (드롭다운). 지연오더는 하단 지연 블록에 모아 적는다'),
    ('B', '라인명',   17, '입력', '라인의 첫 행에만 기재. 2교대는 "라인명/2교대"로 별도 행'),
    ('C', '인원',     6,  '입력', '라인 투입 인원(숫자, 0.5 허용). 라인 첫 행에만'),
    ('D', '잔업',     5,  '입력', '잔업 시 ● (드롭다운). 라인 첫 행에만'),
    ('E', '순서',     5,  '입력', '라인 내 생산 순서. 항상 1부터 (0 시작 금지)'),
    ('F', '오더번호', 11, '입력', '7자리 숫자, 텍스트로 입력(앞자리 0 보존). 한 칸에 하나'),
    ('G', '제품코드', 13, '입력', '텍스트로 입력'),
    ('H', '제품명',   46, '입력', '제품명만. 날짜·오더일·사유를 제품명 뒤에 붙이지 않는다'),
    ('I', '오더일',   9,  '입력', '오더 접수일(날짜). 지연오더는 필수, 정상은 선택'),
    ('J', '오더수량', 10, '입력', '오더 총수량(숫자)'),
    ('K', '생산수량', 10, '입력', '어제까지 누적 생산 실적. 없으면 비움(0)'),
    ('L', '지시수량', 10, '수식', '=오더수량-생산수량 자동. 분할 지시 시 숫자를 직접 덮어쓰면 잔량에 표시'),
    ('M', '잔량',     9,  '수식', '=오더수량-생산수량-지시수량 자동. 0이 아니면 이월 물량'),
    ('N', '비고',     34, '입력', '사유·입고예정일 등. 한 칸에만 적는다(옆 칸 분산 금지)'),
]
LAST = 'N'
FOOTER_LABELS = ['지원1', '지원2·PM', '포장재·세척', '휴가·교육']


# ----------------------------------------------------------------------------- 출력 (통일 양식)
def style_sheet_base(ws):
    ws.sheet_view.showGridLines = False
    for col, hdr, w, *_ in COLS:
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A3'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = '1:2'
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5


def write_plan_sheet(wb, sheet_name, title, rows, footer, note=None):
    """rows: dict 리스트(구분,라인명,인원,잔업,순서,오더번호,제품코드,제품명,오더일,오더수량,생산수량,지시수량,비고)
       footer: {라벨: (인원, 설명)}"""
    ws = wb.create_sheet(sheet_name)
    style_sheet_base(ws)
    # 제목
    ws.merge_cells(f'A1:{LAST}1')
    c = ws['A1']; c.value = title
    c.font = Font(name=FONT, size=16, bold=True, color=NAVY)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    # 헤더
    for col, hdr, *_ in COLS:
        c = ws[f'{col}2']; c.value = hdr
        c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
        c.fill = HDR_FILL; c.border = BORDER
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[2].height = 24

    normal = [r for r in rows if r.get('구분') != '지연']
    delayed = [r for r in rows if r.get('구분') == '지연']
    r = 3
    first_data = r

    def put_row(r, d, delayed=False):
        vals = {
            'A': '지연' if delayed else '정상', 'B': d.get('라인명'), 'C': d.get('인원'),
            'D': d.get('잔업'), 'E': d.get('순서'), 'F': d.get('오더번호'), 'G': d.get('제품코드'),
            'H': d.get('제품명'), 'I': d.get('오더일'), 'J': d.get('오더수량'), 'K': d.get('생산수량'),
            'L': d.get('지시수량') if d.get('지시수량') not in (None, '') else f'=IF(J{r}="","",J{r}-K{r})',
            'M': f'=IF(J{r}="","",J{r}-N(K{r})-N(L{r}))',
            'N': d.get('비고'),
        }
        for col, v in vals.items():
            c = ws[f'{col}{r}']; c.value = v
            c.font = Font(name=FONT, size=10, color='333333')
            c.border = BORDER
            c.alignment = Alignment(horizontal='left' if col in ('H', 'N') else 'center',
                                    vertical='center', wrap_text=True)
            if col in ('J', 'K', 'L', 'M'):
                c.number_format = '#,##0'
            if col in ('L', 'M'):
                c.fill = FORMULA_FILL
            if col == 'I':
                c.number_format = 'mm/dd'
            if col in ('F', 'G') and v is not None:
                c.number_format = '@'
        if d.get('라인명') and r > first_data:      # 라인 구분선
            for col, *_ in COLS:
                cc = ws[f'{col}{r}']
                cc.border = Border(left=thin, right=thin, top=med, bottom=thin)
        ws.row_dimensions[r].height = 20

    for d in normal:
        put_row(r, d); r += 1
    if delayed:
        # 지연 블록 헤더 행
        ws.merge_cells(f'A{r}:{LAST}{r}')
        c = ws[f'A{r}']; c.value = '■ 지연오더 (오더일·사유 필수 기재, 해소 시 정상 라인으로 이동)'
        c.font = Font(name=FONT, size=10, bold=True, color='9C5700'); c.fill = DELAY_FILL
        c.alignment = Alignment(horizontal='left', vertical='center')
        for col, *_ in COLS:
            ws[f'{col}{r}'].border = BORDER
        ws.row_dimensions[r].height = 20
        r += 1
        for i, d in enumerate(delayed, 1):
            d = dict(d); d['라인명'] = d.get('라인명') or '지연오더'; d['순서'] = d.get('순서') or i
            put_row(r, d, delayed=True); r += 1
    last_data = r - 1

    # 하단 지원 블록 (4행 고정) + 총인원
    for label in FOOTER_LABELS:
        cnt, desc = footer.get(label, (None, None))
        ws.merge_cells(f'E{r}:{LAST}{r}')
        vals = {'A': '', 'B': label, 'C': cnt, 'D': None, 'E': desc}
        for col, *_ in COLS:
            c = ws[f'{col}{r}']
            if col in vals: c.value = vals[col]
            c.font = Font(name=FONT, size=10, color='333333'); c.border = BORDER; c.fill = FOOT_FILL
            c.alignment = Alignment(horizontal='left' if col == 'E' else 'center', vertical='center', wrap_text=True)
        ws.row_dimensions[r].height = 20
        r += 1
    ws.merge_cells(f'E{r}:{LAST}{r}')
    ws[f'B{r}'] = '총인원'; ws[f'C{r}'] = f'=SUM(C{first_data}:C{r-1})'
    ws[f'E{r}'] = '라인 인원 + 지원 인원 자동 합계'
    for col, *_ in COLS:
        c = ws[f'{col}{r}']
        c.font = Font(name=FONT, size=10, bold=True, color=NAVY); c.border = BORDER; c.fill = FORMULA_FILL
        c.alignment = Alignment(horizontal='left' if col == 'E' else 'center', vertical='center')
    ws.row_dimensions[r].height = 22
    total_row = r

    # 유효성 검사 / 조건부 서식 (데이터 영역 + 여유 60행)
    rng_end = last_data + 60
    dv1 = DataValidation(type='list', formula1='"정상,지연"', allow_blank=True); dv1.add(f'A3:A{rng_end}')
    dv2 = DataValidation(type='list', formula1='"●"', allow_blank=True); dv2.add(f'D3:D{rng_end}')
    ws.add_data_validation(dv1); ws.add_data_validation(dv2)
    ws.conditional_formatting.add(f'A3:{LAST}{last_data}',
        FormulaRule(formula=['$A3="지연"'], fill=DELAY_FILL))
    ws.conditional_formatting.add(f'M3:M{last_data}',
        FormulaRule(formula=['AND(ISNUMBER($M3),$M3<>0)'], font=Font(name=FONT, size=10, bold=True, color='C00000')))
    ws.print_area = f'A1:{LAST}{total_row}'
    if note:
        ws[f'A{total_row+2}'] = note
        ws[f'A{total_row+2}'].font = Font(name=FONT, size=9, italic=True, color='777777')
    return ws


def write_guide_sheet(wb):
    ws = wb.create_sheet('작성기준', 0)
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 80
    r = 1
    def h(text, size=14):
        nonlocal r
        ws[f'B{r}'] = text; ws[f'B{r}'].font = Font(name=FONT, size=size, bold=True, color=NAVY)
        ws.row_dimensions[r].height = 26; r += 1
    def p(text, color='333333', bold=False):
        nonlocal r
        ws.merge_cells(f'B{r}:E{r}')
        ws[f'B{r}'] = text; ws[f'B{r}'].font = Font(name=FONT, size=10, color=color, bold=bold)
        ws[f'B{r}'].alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[r].height = 18 * max(1, len(text) // 95 + 1); r += 1
    def table(headers, rows, widths=None):
        nonlocal r
        cols = ['B', 'C', 'D', 'E'][:len(headers)]
        for col, hd in zip(cols, headers):
            c = ws[f'{col}{r}']; c.value = hd; c.font = Font(name=FONT, size=10, bold=True, color=NAVY)
            c.fill = HDR_FILL; c.border = BORDER; c.alignment = Alignment(horizontal='center', vertical='center')
        r += 1
        for row in rows:
            for col, v in zip(cols, row):
                c = ws[f'{col}{r}']; c.value = v; c.font = Font(name=FONT, size=10, color='333333')
                c.border = BORDER; c.alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[r].height = 18 * max(1, max(len(str(v)) for v in row) // 70 + 1)
            r += 1
        r += 1

    h('생산계획 일일 계획표 — 통일 작성기준', 16)
    p('치약 / 염모·튜브 / 멀티·직선 세 담당자의 일일 생산계획을 하나의 양식으로 통일한다. '
      '열 구성·위치·이름을 고정해 월별 취합과 지연오더 추출을 사람 손 없이 돌릴 수 있게 하는 것이 목적이다.')
    r += 1
    h('1. 파일 · 시트 · 제목 규칙', 12)
    table(['항목', '기준', '', '예시 / 비고'], [
        ('파일명', '생산계획_{라인군}_{YYYY-MM}.xlsx', '', '생산계획_치약_2026-09.xlsx  (라인군별·월별 1파일, 하루 1시트)'),
        ('시트명', 'MM.DD(요일)', '', '09.04(금)   ※ 한 파일에 여러 라인군을 둘 때만 09.04(금)-염모 처럼 뒤에 붙임'),
        ('제목(A1)', 'MM/DD(요일) 생산계획 - 라인군', '', '09/04(금) 생산계획 - 치약'),
        ('시작 위치', 'A1 제목, 2행 헤더, 3행부터 데이터', '', 'B열부터 시작하거나 헤더가 3행인 양식은 사용하지 않는다'),
        ('글꼴', '맑은 고딕 10pt (제목 16, 헤더 10 굵게)', '', '24pt 등 개별 확대 금지. 크게 보려면 인쇄 배율로 조정'),
        ('인쇄', '가로, 가로 1쪽 맞춤, 1~2행 반복', '', '양식에 이미 설정됨'),
    ])
    h('2. 열 정의 (A~N, 14열 고정 — 열 추가·삭제·순서 변경 금지)', 12)
    table(['열', '항목', '입력/수식', '설명'], [(c, n, k, d) for c, n, w, k, d in COLS])
    h('3. 색상 범례', 12)
    table(['색', '의미', '', '설명'], [
        ('연녹색 헤더', '고정 헤더', '', '수정하지 않는다'),
        ('회색 셀', '수식', '', '지시수량·잔량·총인원. 분할 지시 때만 지시수량에 숫자를 직접 입력'),
        ('연노랑 행', '지연오더', '', '구분이 "지연"이면 자동으로 색이 칠해진다'),
        ('빨간 굵은 숫자', '잔량 ≠ 0', '', '오늘 다 지시하지 못한 이월 물량'),
    ])
    h('4. 작성 규칙', 12)
    for t in [
        '① 순서는 라인마다 1부터. 0으로 시작하지 않는다.',
        '② 라인명·인원·잔업은 라인의 첫 행에만 적고 아래 행은 비운다. 셀 병합은 하지 않는다(병합하면 취합 시 값이 사라진다).',
        '③ 주간 2교대는 "튜브2호/2교대"처럼 라인명 뒤에 붙여 별도 행으로 적는다. "주간2교대"만 단독으로 적지 않는다.',
        '④ 오더번호·제품코드는 한 칸에 하나. 열을 두 개로 나눠 쓰지 않는다(같은 헤더 2개 금지).',
        '⑤ 제품명 뒤에 " / 8/20오더" 같은 날짜나 사유를 붙이지 않는다. 오더일은 I열, 사유는 N열 비고에.',
        '⑥ 비고는 N열 한 칸에만. 옆 칸(O, M 등)에 흘려 쓰지 않는다.',
        '⑦ 지연오더는 정상 라인 아래 "■ 지연오더" 블록에 모으고, 구분=지연, 라인명=지연오더(또는 원래 라인명), 오더일·사유 필수. 해소되면 정상 라인으로 옮긴다.',
        '⑧ 생산수량은 어제까지의 누적 실적. 지시수량은 자동(오더-생산). 오늘 일부만 지시할 때는 지시수량에 숫자를 직접 넣고 잔량을 확인한다.',
        '⑨ 하단 지원 블록은 지원1 / 지원2·PM / 포장재·세척 / 휴가·교육 4행 고정. 인원은 C열 숫자, 담당·이름은 E열에. 총인원은 자동.',
        '⑩ 오더번호·수량은 숫자로만(단위·쉼표·공백 금지). 오더번호는 텍스트 서식이 이미 적용돼 있다.',
    ]:
        p(t)
    r += 1
    h('5. 담당자별로 바뀌는 점', 12)
    table(['담당', '지금', '', '통일 후'], [
        ('치약', 'A열 시작·11열. 지연 블록 없음, 하단 지원1/PM/포장재.세척 3행, 총인원 없음',
         '', '구분·오더일·잔량 3열 추가. 지연오더가 생기면 하단 블록에 기재. 지원 4행 + 총인원 자동'),
        ('염모·튜브', 'B열 시작, 24pt 아리따M/HY견명조, 열 너비 100 이상. 튜브는 오더번호·생산수량 열이 각 2개, '
                     '비고가 L/M/O에 흩어짐, 잔업·인원 셀 병합, "주간2교대" 단독 행, 지연품목 블록 하단',
         '', 'A열 시작·맑은 고딕 10pt. 중복 열 제거, 병합 해제, 비고 N열 하나로. 2교대는 "라인명/2교대". 지연 블록은 하단 표준 블록으로'),
        ('멀티·직선', 'B2 제목·3행 헤더, 지연오더가 맨 위, 순서 0부터, 제품명에 " / 8/20오더" 날짜 삽입, '
                     '생산/지시 수식이 행마다 다름(=I-J 또는 =I-K), #REF! 오류, 지원1/지원2/휴가교육 + SUM',
         '', 'A1 제목·2행 헤더. 지연오더는 하단 블록으로, 순서 1부터, 오더일은 I열로 분리. 지시=오더-생산 단일 수식, 잔량 자동. 지원 4행 표준'),
    ])
    p('※ 예시 시트는 각 담당자의 9/3·9/4 실제 계획을 통일 양식으로 옮긴 것이다. 내용은 그대로, 형식만 바뀌었다.', '777777')
    return ws


def build_template(out):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    write_guide_sheet(wb)
    rows = [
        dict(라인명='치약충전2호', 인원=3, 잔업='●', 순서=1, 오더번호='3826516', 제품코드='310052597',
             제품명='메디안 치석오리지널 치약 120g(BY26', 오더수량=60000, 생산수량=None, 비고=None),
        dict(순서=2, 오더번호='3826228', 제품코드='310052598B', 제품명='메디안 치석화이트 치약 120g(BY26',
             오더수량=48000, 생산수량=8000, 비고='분할 지시 예시: 지시수량에 20000 직접 입력 시 잔량 20000'),
        dict(구분='지연', 오더번호='3820176', 제품코드='766001796A', 제품명='아.리페어 시카 워터풀 500ml(24)',
             오더일=datetime.date(YEAR, 8, 20), 오더수량=4800, 비고='용기 미입고(8/28 미입고) → 9/8 입고예정'),
    ]
    footer = {'지원1': (5, '계획(1) 주임(1) 실적·순회검사(1) 도우미(2)'),
              '지원2·PM': (3, 'PM(3) — 이름 기재'),
              '포장재·세척': (2, '지게차·포장재(1) 세척(1)'),
              '휴가·교육': (0, '')}
    write_plan_sheet(wb, '09.04(금)', '09/04(금) 생산계획 - 라인군', rows, footer,
                     note='※ 예시 행 3개는 지우고 사용. 회색 셀은 수식이므로 지우지 말 것.')
    # 빈 양식 시트 (예시 없이 수식만 30행)
    blank = [dict() for _ in range(30)]
    ws = write_plan_sheet(wb, '양식(빈칸)', 'MM/DD(요일) 생산계획 - 라인군', blank, {})
    wb.save(out)


# ----------------------------------------------------------------------------- 원본 파싱
def _merged_map(ws):
    m = {}
    for rng in ws.merged_cells.ranges:
        anchor = ws.cell(rng.min_row, rng.min_col).value
        for row in range(rng.min_row, rng.max_row + 1):
            for col in range(rng.min_col, rng.max_col + 1):
                m[(row, col)] = anchor
    return m

def _val(ws, mm, r, c):
    v = ws.cell(r, c).value
    if v is None:
        v = mm.get((r, c))
    return v

def _s(v):
    if v is None: return None
    s = str(v).strip()
    return s or None

def _num(v):
    if v is None or v == '': return None
    if isinstance(v, (int, float)): return v
    s = str(v).replace(',', '').strip()
    try: return float(s) if '.' in s else int(s)
    except ValueError: return None

def _seq(v):
    n = _num(v)
    return int(n) if n is not None else None

def _footer_label(s):
    if not s: return None
    t = s.replace(' ', '').replace('.', '').replace('·', '').replace(',', '')
    if t == '지원1': return '지원1'
    if t in ('지원2', 'PM'): return '지원2·PM'
    if t.startswith('포장재'): return '포장재·세척'
    if t.startswith('휴가'): return '휴가·교육'
    return None

def _date_from_title(title):
    m = re.search(r'(\d{1,2})\s*/\s*(\d{1,2})', title) or re.search(r'(\d{1,2})월\s*(\d{1,2})일', title) \
        or re.search(r'(\d{1,2})\.(\d{1,2})', title)
    mo, d = int(m.group(1)), int(m.group(2))
    return datetime.date(YEAR, mo, d)

def _group_from_title(title):
    m = re.search(r'생산계획\s*-\s*(.+)$', title)
    g = (m.group(1) if m else title).strip()
    return {'멀티.크림.턴.팜플': '멀티'}.get(g, g)

ORDER_DATE_RE = re.compile(r'\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})\s*오더\s*$')

def _split_order_date(name):
    if not name: return name, None
    m = ORDER_DATE_RE.search(name)
    if not m: return name.strip(), None
    return name[:m.start()].strip(), datetime.date(YEAR, int(m.group(1)), int(m.group(2)))


def parse_sheet(ws):
    """시트 구조를 판별해 (제목, rows, footer) 반환. 계획표가 아니면 None."""
    mm = _merged_map(ws)
    # 헤더 행/시작 열 찾기
    hdr_row = hdr_col = None
    for r in range(1, 6):
        for c in range(1, 6):
            if _s(ws.cell(r, c).value) == '라인명':
                hdr_row, hdr_col = r, c; break
        if hdr_row: break
    if not hdr_row: return None
    title = _s(ws.cell(hdr_row - 1, hdr_col).value) or _s(ws.cell(1, 1).value) or ws.title
    # 헤더 이름 → 열번호 (중복 헤더는 모두 기억)
    heads = {}
    for c in range(hdr_col, ws.max_column + 1):
        h = _s(_val(ws, mm, hdr_row, c))
        if h:
            key = h.replace(' ', '')
            key = {'제품명(적색:신제품)청색반제품': '제품명', '제품명(적색:신제품)': '제품명'}.get(key, key)
            heads.setdefault(key, []).append(c)
    def col(name): return heads.get(name, [None])[0]
    def get_any(r, name):
        for c in heads.get(name, []):
            v = _val(ws, mm, r, c)
            if v not in (None, '', 0) or (name == '생산수량' and v not in (None, '')):
                if v not in (None, ''): return v
        return None
    line_c, order_c = col('라인명'), col('오더번호')
    # 비고: 헤더 열부터 오른쪽 끝까지 첫 값 (염모 M열, 튜브 O열처럼 흘러간 값 수습)
    remark_c = col('비고')

    rows, footer = [], {}
    cur_line, in_delay, seq_in_line, delay_seq = None, False, 0, 0
    prev_line_name = None
    r = hdr_row + 1
    while r <= ws.max_row:
        line = _s(_val(ws, mm, r, line_c)) if line_c else None
        raw_line = _s(ws.cell(r, line_c).value)
        order = _s(get_any(r, '오더번호'))
        fl = _footer_label(raw_line)
        if fl:
            cnt = _num(ws.cell(r, col('인원')).value) if col('인원') else None
            desc = None
            for c in range(col('인원') + 1, ws.max_column + 1):
                v = _s(ws.cell(r, c).value)
                if v and v != '●': desc = v; break
            footer[fl] = (cnt, desc)
            r += 1; continue
        if raw_line and re.sub(r'\s', '', raw_line) in ('지연오더', '지연품목'):
            in_delay = True; cur_line = None; delay_seq = 0
        elif raw_line and raw_line != '주간2교대' and not in_delay:
            cur_line = raw_line.replace('\n', ' ').replace('/ 교대', '/2교대').replace('/교대', '/2교대')
            seq_in_line = 0
        elif raw_line == '주간2교대':
            base = (prev_line_name or '').split('/')[0]
            cur_line = f'{base}/2교대'; seq_in_line = 0
        elif in_delay and raw_line and not order and not _s(ws.cell(r, col('제품명')).value):
            # 지연 블록 뒤 새 라인명 → 정상 라인 재개 (멀티/직선형은 빈 행으로 끝나므로 여기 안 옴)
            in_delay = False; cur_line = raw_line; seq_in_line = 0
        elif in_delay and raw_line and (not mm.get((r, line_c))):
            in_delay = False; cur_line = raw_line; seq_in_line = 0
        if in_delay and not order and not raw_line and not _s(ws.cell(r, col('제품명')).value):
            # 멀티/직선형: 빈 행이면 지연 블록 종료
            in_delay = False
        if not order and not _s(ws.cell(r, col('제품명')).value):
            # 오더 없는 행 (라인명만 있는 2교대 행 등) → 인원만 있으면 인원 행으로 보존
            cnt = _num(ws.cell(r, col('인원')).value) if col('인원') else None
            if raw_line and cnt is not None and raw_line == '주간2교대':
                rows.append(dict(구분='정상', 라인명=cur_line, 인원=cnt, 잔업=_s(ws.cell(r, col('잔업')).value)))
                prev_line_name = cur_line
            r += 1; continue
        name, odate = _split_order_date(_s(_val(ws, mm, r, col('제품명'))))
        first_of_line = bool(raw_line) or (in_delay and False)
        if in_delay:
            delay_seq += 1
        else:
            seq_in_line += 1
        d = dict(
            구분='지연' if in_delay else '정상',
            라인명=(cur_line if first_of_line else None) if not in_delay else None,
            인원=_num(ws.cell(r, col('인원')).value) if col('인원') else None,
            잔업=(_s(ws.cell(r, col('잔업')).value) or (mm.get((r, col('잔업'))) if first_of_line else None)) if col('잔업') else None,
            순서=delay_seq if in_delay else seq_in_line,
            오더번호=order, 제품코드=_s(get_any(r, '제품코드')), 제품명=name, 오더일=odate,
            오더수량=_num(get_any(r, '오더수량')),
            생산수량=None, 지시수량=None,
            비고=None,
        )
        # 생산/지시: 수식이면 값으로 환산 (I-J / I-K 형태)
        prod = get_any(r, '생산수량'); inst = get_any(r, '지시수량')
        def ev(v):
            if isinstance(v, str) and v.startswith('='):
                m = re.fullmatch(r'=([A-Z]+)(\d+)-([A-Z]+)(\d+)', v.replace(' ', ''))
                if m:
                    a = _num(ws[f'{m.group(1)}{m.group(2)}'].value); b = _num(ws[f'{m.group(3)}{m.group(4)}'].value)
                    return (a or 0) - (b or 0)
                return None
            return _num(v)
        prod_v, inst_v = ev(prod), ev(inst)
        if prod_v is None and inst_v is not None and d['오더수량'] is not None and inst_v != d['오더수량']:
            prod_v = d['오더수량'] - inst_v
        d['생산수량'] = prod_v if prod_v else None
        # 지시가 오더-생산과 다르면(분할지시) 직접 값 유지
        if inst_v is not None and d['오더수량'] is not None and inst_v != d['오더수량'] - (prod_v or 0):
            d['지시수량'] = inst_v
        # 비고: 비고 헤더 열 ~ 끝까지 첫 텍스트
        if remark_c:
            for c in range(remark_c, ws.max_column + 1):
                v = _s(ws.cell(r, c).value)
                if v: d['비고'] = v.lstrip("'"); break
        if d['잔업'] and d['잔업'] != '●': d['잔업'] = '●'
        rows.append(d); prev_line_name = cur_line or prev_line_name
        r += 1
    # 지연 블록: 라인명 없이 모아서 기재 (원래 라인명이 있으면 유지)
    return title, rows, footer


def convert(out, sources, sheets=None):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    for src in sources:
        swb = openpyxl.load_workbook(src)
        for ws in swb.worksheets:
            if sheets and ws.title not in sheets: continue
            parsed = parse_sheet(ws)
            if not parsed: continue
            title, rows, footer = parsed
            d = _date_from_title(title); g = _group_from_title(title)
            t = f'{d.month:02d}/{d.day:02d}({WD[d.weekday()]}) 생산계획 - {g}'
            name = f'{d.month:02d}.{d.day:02d}({WD[d.weekday()]})-{g}'
            write_plan_sheet(wb, name, t, rows, footer)
            print(f'{src.split("/")[-1]} / {ws.title} -> {name}: {len(rows)}행, 지연 {sum(1 for x in rows if x["구분"]=="지연")}행')
    wb.save(out)


if __name__ == '__main__':
    cmd = sys.argv[1]
    if cmd == 'template':
        build_template(sys.argv[2])
    elif cmd == 'convert':
        out = sys.argv[2]; srcs = [a for a in sys.argv[3:] if not a.startswith('--sheets=')]
        sh = [a for a in sys.argv[3:] if a.startswith('--sheets=')]
        convert(out, srcs, sh[0][9:].split(',') if sh else None)
